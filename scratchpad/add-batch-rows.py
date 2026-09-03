"""
add-batch-rows.py — append empty workbook rows for a new batch, keyed by filename.

The workbook is a flat per-image sheet keyed by `filename`, and a new batch has no
rows at all. Rather than have the user create 99 rows by hand, this appends them with
`filename` and `itemID` filled and everything else blank, ready to type into.

Harmless if left half-filled: excel_to_json.py only updates records that already exist
in data/museum-data.json, so a blank row creates nothing and changes nothing. The rows
become live only when the records themselves are built.

  py scratchpad/add-batch-rows.py                 # dry run
  py scratchpad/add-batch-rows.py --write
"""
import json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

src = json.load(open('scratchpad/new-batch-paths.json', encoding='utf-8'))
ids = sorted('goetzmann' + n for n in src)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

existing = set()
last = 1
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip():
        existing.add(str(v).strip().lower())
        last = r

todo = [i for i in ids if (i + '.jpg').lower() not in existing]
print('%d ids in the batch; %d already have a row; %d to append' % (len(ids), len(ids) - len(todo), len(todo)))
if not todo:
    sys.exit(0)
print('  appending rows %d-%d for %s … %s' % (last + 1, last + len(todo), todo[0], todo[-1]))
print('  columns filled: filename, itemID (blank elsewhere)')
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

# itemID: continue the sheet's own numbering if it is numeric, else leave blank
nums = []
for r in range(2, last + 1):
    v = ws.cell(row=r, column=head['itemID']).value if 'itemID' in head else None
    if v is not None and str(v).strip().isdigit():
        nums.append(int(str(v).strip()))
nxt = max(nums) + 1 if nums else None

tmp = BOOK + '.tmp'
row = last
for i, rid in enumerate(todo):
    row += 1
    ws.cell(row=row, column=head['filename'], value=rid + '.jpg')
    if nxt is not None and 'itemID' in head:
        ws.cell(row=row, column=head['itemID'], value=nxt + i)
wb.save(tmp)

before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if len(after) != len(before) + len(todo):
    os.remove(tmp); sys.exit('ABORT: row count %d -> %d, expected +%d' % (len(before), len(after), len(todo)))
# every pre-existing cell must be untouched
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns if before.at[i, c] != after.at[i, c]]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d existing cell(s) changed: %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-batchrows')
os.replace(tmp, BOOK)
print('\nverified: %d rows appended, 0 existing cells touched' % len(todo))

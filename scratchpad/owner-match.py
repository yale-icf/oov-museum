"""
owner-match.py — match the new workbook's owner data onto our records.

Matches on the image name (Filename -> record id) and then CONFIRMS each match on
the description, as asked. A confirmation step is needed because a filename alone
proves nothing if files were ever renumbered, and our descriptions were rewritten
from scratch this month, so they will never match the sheet's textually.

The confirmation therefore compares DISTINCTIVE tokens rather than prose: proper
nouns, 4-digit years, and long numbers that appear in both sides' title +
description. Two records about the same object share those; two unrelated ones do
not.

Source sheet: "Goetzmann File Names". It is the updated one -- where the other
sheets say flatly "Beinecke" it distinguishes ICF, joint ICF/Beinecke purchases and
transfers, and it covers 467 of our 474 records against their 455.

  py scratchpad/owner-match.py            # report only
  py scratchpad/owner-match.py --write    # write owner to the workbook + JSON
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

SRC = 'virtual_museum_database_full.xlsx'
BOOK = 'oov_data_new_edit_2.xlsx'
SHEET = 'Goetzmann File Names'
WRITE = '--write' in sys.argv

if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

# Matched on filename but scoring low, because the sheet's text for them is terse or in
# another language -- "Kaiserlich Russische Regierung", "Namkok Gold Mine Company",
# "Emprunt industriel du gouvernement de la Republique Chinoise", "Kriegsanleihe". Each was
# read against our title by eye on 2026-08-28 and is unmistakably the same object, so they
# are admitted explicitly rather than by loosening the threshold for everything.
# The four Beinecke provenance variants collapse to plain "Beinecke" -- user decision
# 2026-08-28. Applied on read so this script is idempotent and cannot undo it.
COLLAPSE = {
    'Beinecke (Purchased by ICF transferred to Beinecke)': 'Beinecke',
    'Part of DLJ Collection - Joint purchase with ICF/Beinecke': 'Beinecke',
    'Joint ICF Beinecke Purchase': 'Beinecke',
    'Purchased by ICF transferred to Beinecke in 2003': 'Beinecke',
}

VERIFIED_WEAK = {
    'goetzmann0218', 'goetzmann0226', 'goetzmann0314', 'goetzmann0324', 'goetzmann0343',
    'goetzmann0361', 'goetzmann0387', 'goetzmann0390', 'goetzmann0399', 'goetzmann0412',
    'goetzmann0419', 'goetzmann0435', 'goetzmann0445', 'goetzmann0463', 'goetzmann0467',
    'goetzmann0518', 'goetzmann0519', 'goetzmann0572', 'goetzmann0589',
}

STOP = set(('the a an of and for in on at by to with from this that is was were be been it its '
            'document page one two three bond share note certificate loan company government '
            'state national bank issued printed dated paper front back reverse recto verso '
            'collection goetzmann jpg image scan file').split())

def tokens(*parts):
    """distinctive tokens: capitalised words 4+ letters, years, and long numbers"""
    s = ' '.join(p or '' for p in parts)
    out = set()
    for w in re.findall(r"\b[A-ZÀ-Þ][A-Za-zÀ-ÿ'’-]{3,}\b", s):
        w = w.lower().strip("'’-")
        if w not in STOP:
            out.add(w)
    out |= set(re.findall(r'\b1[0-9]{3}\b|\b20[0-2][0-9]\b', s))
    out |= set(n.replace(',', '') for n in re.findall(r'\b\d[\d,]{3,}\b', s))
    return out

# ---- load the source sheet ----
wbs = load_workbook(SRC, read_only=True, data_only=True)
ws = wbs[SHEET]
rows = ws.iter_rows(values_only=True)
hdr = list(next(rows))
idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
get = lambda r, k: (r[idx[k]] if k in idx and idx[k] < len(r) else None)

src = {}
for r in rows:
    f = get(r, 'Filename')
    if not f:
        continue
    k = str(f).strip()
    if not k.lower().startswith('goetzmann'):
        continue
    # ⚠️ 12 rows spell the id with a capital G ("Goetzmann0900"); lowercase the KEY,
    # not just the filter, or those records silently miss out. Seven of ours did.
    rid = (k[:-4] if k.lower().endswith('.jpg') else k).lower()
    # ten cells end in a stray " = ", a spreadsheet artifact, and five hold a lone space
    owner = re.sub(r'[\s=]+$', '', str(get(r, 'Owner') or '')).strip()
    owner = COLLAPSE.get(owner, owner)
    src[rid] = {
        'owner': owner,
        'tok': tokens(get(r, 'Document Title'), get(r, 'Description')),
    }

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

# ---- match and confirm ----
matched, weak, missing, noowner = [], [], [], []
for rid, rec in sorted(by_id.items()):
    if rid not in src:
        missing.append(rid); continue
    s = src[rid]
    if not s['owner']:
        noowner.append(rid); continue
    ours = tokens(rec.get('title'), rec.get('description'))
    shared = ours & s['tok']
    denom = min(len(ours), len(s['tok'])) or 1
    score = len(shared) / denom
    row = (rid, s['owner'], len(shared), round(score, 2), sorted(shared)[:6])
    ok = len(shared) >= 3 or score >= 0.25 or rid in VERIFIED_WEAK
    (matched if ok else weak).append(row)

print('%d records; %d have no row in the sheet; %d have a row but a blank owner'
      % (len(by_id), len(missing), len(noowner)))
print('%d confirmed by description, %d matched by filename but NOT confirmed\n' % (len(matched), len(weak)))

from collections import Counter
print('owner values to apply (confirmed only):')
for v, k in Counter(m[1] for m in matched).most_common():
    print('  %4d  %s' % (k, v))

print('\nweak matches -- filename lines up, description does not (%d):' % len(weak))
for rid, owner, n, sc, sh in weak[:20]:
    print('  %s  owner=%-12s shared=%d  %s' % (rid, owner[:12], n, sh))
    print('       ours : ' + (by_id[rid].get('title') or '')[:82])

if missing:
    print('\nno row in the sheet (%d): %s' % (len(missing), ', '.join(missing[:12])))

# ---- differences against what we already hold ----
changes = [(rid, (by_id[rid].get('owner') or '').strip(), owner)
           for rid, owner, n, sc, sh in matched
           if (by_id[rid].get('owner') or '').strip() != owner]
conflicts = [c for c in changes if c[1]]
print('\n%d records would gain or change an owner; %d of those already hold a DIFFERENT value:'
      % (len(changes), len(conflicts)))
for rid, old, new in conflicts:
    print('  %s  %r -> %r' % (rid, old, new))

if not WRITE:
    print('\nreport only -- pass --write to apply')
    sys.exit(0)

# ---- write ----
wb = load_workbook(BOOK)
wt = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(wt.cell(row=1, column=c).value).strip(): c
        for c in range(1, wt.max_column + 1) if wt.cell(row=1, column=c).value}
wrows = {}
for r in range(2, wt.max_row + 1):
    v = wt.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        wrows[str(v).strip()[:-4]] = r

edits = []
for rid, old, new in changes:
    if rid not in wrows:
        sys.exit('no workbook row for ' + rid)
    edits.append((rid, wrows[rid], head['owner'], old, new))
tmp = BOOK + '.tmp'
for rid, r, c, old, new in edits:
    wt.cell(row=r, column=c, value=new)
wb.save(tmp)

before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, 'owner') for _, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-owner')
os.replace(tmp, BOOK)

for rid, r, c, old, new in edits:
    by_id[rid]['owner'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; %d owner cells written to the workbook and the JSON' % len(edits))

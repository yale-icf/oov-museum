"""
owner-followup.py — the corrections that came out of the user's question
"ICF only has 29 pieces?", 2026-08-28.

1. CASE BUG in my first import. Twelve rows spell the id with a capital G,
   "Goetzmann0900.jpg". I lowercased only for the filter, not for the key, so those
   never matched. Seven were ours and ALL SEVEN are ICF. Their titles confirm the
   identity exactly -- Silesian Loan of Karl VI, Commerce d'Asie & d'Afrique
   (Trieste Company), the Caracas Company, Swedish West-India Company -- and
   0900/0902/0903 sit under "Stadtanleihen Mecklenburg-Vorpommern", which is right:
   Hagenow and Rostock are both in Mecklenburg.

   Five of the seven already carried a hand-entered owner that the database
   contradicts. The database is authoritative, so ICF wins.

2. goetzmann1002 -> WNG. The user's rule was "blank owner but status Found can be
   ICF". Only one blank-owner record has status Found, and it is this one -- but its
   own second leaf, goetzmann1003, the same document, is explicitly WNG. Leaf
   evidence beats a default, so WNG is written and the departure is reported.

  py scratchpad/owner-followup.py            # dry run
  py scratchpad/owner-followup.py --write
"""
import io, json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

PLAN = {
    'goetzmann0900': 'ICF', 'goetzmann0902': 'ICF', 'goetzmann0903': 'ICF',
    'goetzmann0904': 'ICF', 'goetzmann0908': 'ICF', 'goetzmann0909': 'ICF',
    'goetzmann0910': 'ICF',
    'goetzmann1002': 'WNG',
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
col = head['owner']
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

edits = []
for rid, new in PLAN.items():
    r = rows[rid]
    cur = '' if ws.cell(row=r, column=col).value is None else str(ws.cell(row=r, column=col).value).strip()
    if cur != new:
        edits.append((rid, r, cur, new))

print('%d owner cell(s) to write' % len(edits))
for rid, r, old, new in edits:
    print('  %s  %-9s -> %-4s   %s' % (rid, repr(old), new, (by_id[rid].get('title') or '')[:56]))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
for rid, r, old, new in edits:
    ws.cell(row=r, column=col, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, 'owner') for _, r, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-ownerfollowup')
os.replace(tmp, BOOK)
for rid, r, old, new in edits:
    by_id[rid]['owner'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; %d cells written to the workbook and the JSON' % len(edits))

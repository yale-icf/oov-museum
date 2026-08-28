"""
apply-titles.py

Writes the uniform titles into the workbook and data/museum-data.json together.
`title` is Excel-backed, so the sheet has to carry them or the next import reverts.

Rules applied (see scratchpad/title-uniform.js): no issue date, no interest rate,
no denomination, serial, series or class. Years that are part of a name --
"Gold Loan of 1907", "German External Loan 1924", "8 George I (1722)" -- survive.

Safety: works on a copy of the workbook, verifies every other cell is untouched,
keeps a .bak, refuses to run against an open workbook.

  py scratchpad/apply-titles.py                 # dry run
  py scratchpad/apply-titles.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

new = json.load(open('scratchpad/title-rewrites.json', encoding='utf-8'))
recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()
cell = lambda r, c: '' if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value)

edits, skipped = [], []
for rid, title in sorted(new.items()):
    if rid not in rows:
        sys.exit('no workbook row for ' + rid)
    r, c = rows[rid], head['title']
    # the sheet must still hold the title the proposal was computed from
    if norm(cell(r, c)) != norm(by_id[rid].get('title')):
        skipped.append(rid); continue
    if norm(cell(r, c)) != norm(title):
        edits.append((rid, r, c, cell(r, c), title))

if skipped:
    print('SKIPPED %d row(s) where the sheet and JSON disagree: %s' % (len(skipped), ', '.join(skipped)))
print('%d title cells to write' % len(edits))
for rid, r, c, old, t in edits[:6]:
    print('  %s  row %d\n     %s\n  -> %s' % (rid, r, old, t))
if len(edits) > 6:
    print('  … %d more (full list in scratchpad/title-diff.tsv)' % (len(edits) - 6))

if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, r, c, old, t in edits:
    ws.cell(row=r, column=c, value=t)
wb.save(tmp)

before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if before.shape != after.shape:
    os.remove(tmp); sys.exit('ABORT: shape changed')
intended = {(r - 2, 'title') for _, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
changed = sum(1 for i in range(len(before)) for col in before.columns
              if before.at[i, col] != after.at[i, col])
print('\nverified: %d cells changed, %d intended, 0 collateral' % (changed, len(edits)))

shutil.copy2(BOOK, BOOK + '.bak-before-titles')
os.replace(tmp, BOOK)

for rid, r, c, old, t in edits:
    by_id[rid]['title'] = t
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('WRITTEN to the workbook and data/museum-data.json')

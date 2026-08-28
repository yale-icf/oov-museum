"""
provenance-import.py — bring the acquisition columns across from the virtual museum
database, which the user has confirmed holds the correct information.

Imports into oov_data_new_edit_2.xlsx:  purchaseDate, purchasePrice, purchasedFrom.
All three are entirely empty in our workbook (0 of 474) and are not read by
excel_to_json.py at all, so they never reach data/museum-data.json and are never
rendered. The workbook is their proper home.

⚠️ Deliberately NOT imported, though the sheet carries them:

  Type      the sheet holds the old free text with _x001D_ separators --
            "Bond_x001D_debt_x001D_security". Our `type` is a controlled vocabulary
            of 21 values settled in July 2026, and the browse shelves in
            museum-search.js are built on it. Overwriting would break browse and
            reintroduce the separator artifacts.
  Period    the sheet has values like "18th Century or before", which is not our
            vocabulary (Pre-17th Century / 17th / 18th / 19th / 20th / 21st).
  Location  ours is DERIVED by excel_to_json.py from subjectCountry + issuingCountry
            and was repaired at the August import; it is not a free-text field.
  Keywords  the sheet's keywords are filename-stuffed and carry _x001D_ artifacts.

Those four are curated fields where our data is the later and better version. They
are reported, not written.

  py scratchpad/provenance-import.py            # report only
  py scratchpad/provenance-import.py --write
"""
import datetime as dt
import json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

SRC = 'virtual_museum_database_full.xlsx'
BOOK = 'oov_data_new_edit_2.xlsx'
SHEET = 'Goetzmann File Names'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)
if os.path.exists('~$' + SRC):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % SRC)

FIELDS = {'Purchase Date': 'purchaseDate',
          'Purchase Price': 'purchasePrice',
          'Purchased From': 'purchasedFrom'}

def clean(v):
    if v is None:
        return ''
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime('%Y-%m-%d')
    return re.sub(r'[\s=]+$', '', str(v).replace('_x001D_', '; ')).strip()

wbs = load_workbook(SRC, read_only=True, data_only=True)
ws = wbs[SHEET]
rows = ws.iter_rows(values_only=True)
hdr = list(next(rows))
idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
src = {}
for r in rows:
    f = r[idx['Filename']] if idx['Filename'] < len(r) else None
    if not f or not str(f).lower().startswith('goetzmann'):
        continue
    rid = str(f).strip()
    rid = rid[:-4] if rid.lower().endswith('.jpg') else rid
    src[rid] = {ours: clean(r[idx[col]] if col in idx and idx[col] < len(r) else None)
                for col, ours in FIELDS.items()}

recs = {r['id'] for r in json.load(open('data/museum-data.json', encoding='utf-8'))}

wb = load_workbook(BOOK)
wt = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(wt.cell(row=1, column=c).value).strip(): c
        for c in range(1, wt.max_column + 1) if wt.cell(row=1, column=c).value}
for f in FIELDS.values():
    if f not in head:
        sys.exit('no %r column in %s' % (f, BOOK))

edits = []
for r in range(2, wt.max_row + 1):
    v = wt.cell(row=r, column=head['filename']).value
    if not v or not str(v).strip().endswith('.jpg'):
        continue
    rid = str(v).strip()[:-4]
    if rid not in recs or rid not in src:
        continue
    for field, value in src[rid].items():
        if not value:
            continue
        c = head[field]
        cur = '' if wt.cell(row=r, column=c).value is None else str(wt.cell(row=r, column=c).value).strip()
        if cur != value:
            edits.append((rid, field, r, c, cur, value))

from collections import Counter
print('%d cell(s) to write across %d record(s)'
      % (len(edits), len({e[0] for e in edits})))
for f in FIELDS.values():
    vals = [e[5] for e in edits if e[1] == f]
    print('\n  %s — %d' % (f, len(vals)))
    for v, k in Counter(vals).most_common(6):
        print('      %3d  %s' % (k, v[:78]))

if not WRITE:
    print('\nreport only -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    wt.cell(row=r, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-provenance')
os.replace(tmp, BOOK)
print('\nverified 0 collateral; %d cells written to %s' % (len(edits), BOOK))
print('These columns are workbook-only: excel_to_json.py does not read them, so')
print('data/museum-data.json is unchanged and the site is unaffected.')

"""
import-batch.py — create the 47 new records from To_Add_Files.xlsx.

Imports only what the source actually has: title (VERBATIM — the user is revamping
titles separately), description, owner, and period where it is already one of our
values. Everything else is deliberately left empty for the cataloguing pass:

  type          the source has 2 of 99, in the old free-text form
                ("security_x001D_debt"), which would break the browse shelves
  period        "18th Century or before" is NOT our vocabulary — it collapses 17th
                and Pre-17th. Those 24 rows are left BLANK rather than guessed, so
                no phantom facet value is created
  issueDate     no such column in the source, so no issueYear, so no Date row
  issuingCountry / subjectCountry / currency / language / keywords / identifiers
                absent from the source

The one edit made to the text: the leading "Page 1 - " / "Page 5 (Insert 1 front
side) -" marker is stripped. That is structural metadata which the pages[] array now
carries, not content.

Writes the records to data/museum-data.json AND the rows to the workbook, so the
workbook stays the source of truth and the user can edit there.

  py scratchpad/import-batch.py                 # dry run
  py scratchpad/import-batch.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

VALID_PERIOD = {'Pre-17th Century', '17th Century', '18th Century', '19th Century',
                '20th Century', '21st Century'}
PAGE = re.compile(r'^\s*Page\s*\d+\s*(?:of\s*\d+)?\s*(?:\([^)]*\))?\s*[-–:]?\s*', re.I)

rows = json.load(open('scratchpad/toadd-rows.json', encoding='utf-8'))
hdr = json.load(open('scratchpad/toadd-header.json', encoding='utf-8'))
idx = {h: i for i, h in enumerate(hdr) if h}
groups = json.load(open('scratchpad/batch-groups.json', encoding='utf-8'))

def cell(rid, k):
    v = rows[rid][idx[k]] if k in idx and idx[k] < len(rows[rid]) else None
    return '' if v in (None, 'None') else re.sub(r'\s+', ' ', str(v)).strip()

def body(rid):
    return PAGE.sub('', cell(rid, 'Description')).strip()

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
have = {r['id'] for r in recs}

new, skipped = [], []
for g in groups:
    rid = g[0]
    if rid in have:
        skipped.append(rid); continue
    per = cell(rid, 'Period')
    rec = {
        'id': rid,
        'title': cell(rid, 'Document Title'),
        'description': body(rid),
        'type': [],
        'location': [],
        'subjectCountry': [],
        'issuingCountry': [],
        'creator': '',
        'issueYear': [],
        'currency': [],
        'language': [],
        'period': [per] if per in VALID_PERIOD else [],
        'keywords': [],
        'owner': cell(rid, 'Owner'),
        'notes': '',
        'identifiers': [],
        'namedIndividuals': [],
        'transcription': '',
    }
    if len(g) > 1:
        rec['pages'] = [{'id': x, 'description': body(x)} for x in g]
    new.append(rec)

print('%d documents to add (%d skipped, already present)' % (len(new), len(skipped)))
print('  multi-leaf: %d   single: %d'
      % (sum(1 for r in new if r.get('pages')), sum(1 for r in new if not r.get('pages'))))
print('  with a usable period: %d of %d' % (sum(1 for r in new if r['period']), len(new)))
print('  collection: %d -> %d' % (len(recs), len(recs) + len(new)))
print('\nfirst three:')
for r in new[:3]:
    print('   %s  %s' % (r['id'], r['title'][:66]))
    print('        %s' % r['description'][:100])
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

# ---- JSON ----
recs.extend(new)
recs.sort(key=lambda r: r['id'])
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nJSON: %d records' % len(recs))

# ---- workbook: one row per IMAGE ----
wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
existing, last = set(), 1
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip():
        existing.add(str(v).strip().lower()); last = r

order = [rid for g in groups for rid in g]
todo = [i for i in order if (i + '.jpg').lower() not in existing]
primaries = {g[0] for g in groups}
tmp = BOOK + '.tmp'
row = last
for rid in todo:
    row += 1
    ws.cell(row=row, column=head['filename'], value=rid + '.jpg')
    ws.cell(row=row, column=head['title'], value=cell(rid, 'Document Title'))
    ws.cell(row=row, column=head['description'], value=body(rid))
    ws.cell(row=row, column=head['owner'], value=cell(rid, 'Owner'))
    per = cell(rid, 'Period')
    if per in VALID_PERIOD:
        ws.cell(row=row, column=head['period'], value=per)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if len(after) != len(before) + len(todo):
    os.remove(tmp); sys.exit('ABORT: rows %d -> %d, expected +%d' % (len(before), len(after), len(todo)))
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns if before.at[i, c] != after.at[i, c]]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d existing cell(s) changed: %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-batchimport')
os.replace(tmp, BOOK)
print('workbook: %d rows appended, 0 existing cells touched' % len(todo))

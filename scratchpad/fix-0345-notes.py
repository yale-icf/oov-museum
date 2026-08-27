"""
fix-0345-notes.py

goetzmann0345's `notes` carry two errors documented in docs/PENDING_XLSX_FIXES.md,
both verified from the image (scratchpad/reconstruct.js) 2026-08-27:

  "No. DP 1602"            the "DP" is an ornate No. glyph misread by OCR. The
                           coupons are stamped a plain 1602, and the record's
                           `identifiers` already reads "No. 1602".
  "principal due July 1876" the docket reads JULY 1893. 1876 is the Redeemer year,
                           conflated in from the old description.

The user has already rewritten the description in the sheet and it is correct
(July 1893, and it no longer calls the scan "the engraved face" -- it is the back).
Only notes was left carrying the errors, identically in both the sheet and the JSON,
so an import would not have caught it.

  py scratchpad/fix-0345-notes.py                 # dry run
  py scratchpad/fix-0345-notes.py --write
"""
import io, json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

OLD = ('State of South Carolina, Six Per Cent Consolidation Bond, $500, No. DP 1602, '
       'secured by annual tax, principal due July 1876')
NEW = ('State of South Carolina, Six Per Cent Consolidation Bond, $500, No. 1602, '
       'secured by annual tax, principal due July 1893')

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == 'goetzmann0345.jpg')
cur = str(ws.cell(row=row, column=head['notes']).value or '').strip()
if cur != OLD:
    sys.exit('sheet notes are not what was expected:\n  %r' % cur)

print('row %d notes\n  was: %s\n  now: %s' % (row, OLD, NEW))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
ws.cell(row=row, column=head['notes'], value=NEW)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) != (row - 2, 'notes')]
if bad:
    os.remove(tmp); sys.exit('ABORT: unintended changes %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-0345')
os.replace(tmp, BOOK)

p = 'data/museum-data.json'
d = json.load(open(p, encoding='utf-8'))
next(x for x in d if x['id'] == 'goetzmann0345')['notes'] = NEW
io.open(p, 'w', encoding='utf-8', newline='\n').write(json.dumps(d, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

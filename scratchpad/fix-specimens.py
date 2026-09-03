"""
fix-specimens.py — distinguish the specimen sets by their ID numbers.

Three duplicate-title groups came in with the 1100-1199 batch, all specimen sets where
each item is a genuinely different certificate sharing a series title. Each carries its
own ID number in the description, so that is the distinguisher — the same convention
used for goetzmann0386/0669 ("..., No. 116485").

The ID also goes into `identifiers`, which the search page indexes. That matters
because the titles are due for a revamp: the number will still separate these records
even after the title text changes.

  py scratchpad/fix-specimens.py                 # dry run
  py scratchpad/fix-specimens.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

IDS = {
    'goetzmann1124': '85-NF-0000000',
    'goetzmann1125': '86-ND-0000000',
    'goetzmann1126': '86-NX-0000000',
    'goetzmann1127': '85-NX-0000000',
    'goetzmann1128': '88-NG-0000000',
    'goetzmann1129': '88-NF-0000000',
    'goetzmann1130': '88-NX-0000000',
    'goetzmann1131': '88-ND-0000000',
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

# every ID must actually appear in that record's own description
for rid, num in IDS.items():
    if num not in norm(by_id[rid].get('description')):
        sys.exit('%s: %s is not in its description' % (rid, num))
if len(set(IDS.values())) != len(IDS):
    sys.exit('ABORT: the ID numbers are not unique')

plan = {}
for rid, num in IDS.items():
    t = norm(by_id[rid]['title'])
    t = re.sub(r',\s*No\.\s*\S+$', '', t)          # idempotent
    plan[rid] = {'title': '%s, No. %s' % (t, num), 'identifiers': 'No. ' + num}

print('%d specimen titles to write' % len(plan))
for rid, f in sorted(plan.items()):
    print('   %s  %s' % (rid, f['title']))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

edits = []
for rid, fields in plan.items():
    row = rows[rid]
    for f, v in fields.items():
        c = head[f]
        cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
        if norm(cur) != norm(v):
            edits.append((rid, f, row, c, v))
tmp = BOOK + '.tmp'
for rid, f, row, c, v in edits:
    ws.cell(row=row, column=c, value=v)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-specimens')
os.replace(tmp, BOOK)

for rid, fields in plan.items():
    by_id[rid]['title'] = fields['title']
    by_id[rid]['identifiers'] = [fields['identifiers']]
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; %d cells written to the workbook and the JSON' % len(edits))

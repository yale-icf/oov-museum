"""
restore-duplicates.py — bring back the two genuine duplicate records, 2026-09-02.

The user asked for the removed duplicates back, presented identically: the same
title, distinguished only by year or serial, and one shared description per pair.

⚠️ Only TWO of the four are genuine duplicates. Read from the images:

  0493 / 0494  the SAME physical debenture photographed twice — both No. 291, both
               George Crossman of 13 Drury Lane, 18 June 1897, same signatures, same
               CANCELLED stamp in the same position, same folds, punch holes and
               pencil "1509". Only the exposure differs.
  1031 / 1027  the SAME document photographed twice — same "1627 11 Genn°" note,
               same "quattro e mezzo" luoghi, same "quattrocento cinquanta" scudi,
               same signatures, same wax seal in the same spot. 1027 is shot on
               black, 1031 on white.

For those two there is no second object, so no year and no serial CAN differ — the
user's rule cannot be satisfied. They are left out and reported.

Restored here:

  0294 / 0295  Banca Românească, two emissions — 1938 and 1920. Distinguished by
               YEAR. Genuinely different: 350,000,000 lei of subscribed capital
               against 160,000,000.
  0386 / 0669  Russian 4% State Loan of 1902, two bonds of the same issue.
               Distinguished by SERIAL, No. 116485 and No. 196422, since both are
               1902.

Shared descriptions:

  the Russian pair takes 0669's text unchanged — it is the fuller of the two and
  nothing in it is specific to one copy;
  the Banca pair needs a new shared text, since the originals differ on year,
  capital, emission and historical framing. What both say — 500 lei, bearer,
  Bucharest, the bank, and the same allegorical vignettes — is kept; what is true of
  only one is dropped and now lives in the title's year.

  py scratchpad/restore-duplicates.py                 # dry run
  py scratchpad/restore-duplicates.py --write
"""
import io, json, os, re, shutil, subprocess, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

BANCA = ("A bearer share of 500 lei in Banca Românească, one of Romania's principal commercial "
         "banks, founded in 1911, issued at Bucharest. Because it is a bearer share, whoever held "
         "the certificate owned it, with no name registered. Allegorical vignettes tie the bank's "
         "promise to the sources of national wealth: a standing labourer with his tool, a seated "
         "figure amid the fruits of the harvest, and a steamer crossing the lower border, binding "
         "labour, agriculture and seaborne trade to the bank's prospects. The dense engraved "
         "guilloche border served both as ornament and as a guard against counterfeiting.")

TITLES = {
    'goetzmann0294': 'Banca Românească Share Certificate, 1938',
    'goetzmann0295': 'Banca Românească Share Certificate, 1920',
    'goetzmann0386': 'Imperial Russian Government State Loan of 1902 Bond, No. 116485',
    'goetzmann0669': 'Imperial Russian Government State Loan of 1902 Bond, No. 196422',
}
RESTORE = ['goetzmann0294', 'goetzmann0386']
NOTES = {
    'goetzmann0294': ('Banca Românească, 500 lei share, Bucharest 1938, eighth emission (Emisiunea '
                      'VIII-A), subscribed capital 350,000,000 lei. Same engraved plate as the 1920 '
                      'share goetzmann0295; the pair is shown with a shared description, the year in '
                      'the title carrying the difference. Restored to the site 2026-09-02.'),
    'goetzmann0386': ('Russian 4% State Loan of 1902, 1,000 German marks, No. 116485. The same bond '
                      'as goetzmann0669 (No. 196422); the pair is shown with a shared description, '
                      'the serial in the title carrying the difference. Restored to the site '
                      '2026-09-02.'),
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}
recovered = json.load(open('scratchpad/restore-records.json', encoding='utf-8'))

# ---- build the restored records ----
added = []
for rid in RESTORE:
    if rid in by_id:
        print('%s is already present, skipping' % rid); continue
    r = dict(recovered[rid])
    r['title'] = TITLES[rid]
    r['description'] = BANCA if rid == 'goetzmann0294' else by_id['goetzmann0669']['description']
    r['notes'] = NOTES[rid]
    added.append(r)

# twins: shared description + title
twin_updates = {
    'goetzmann0295': {'title': TITLES['goetzmann0295'], 'description': BANCA},
    'goetzmann0669': {'title': TITLES['goetzmann0669']},
}

print('restoring %d record(s): %s' % (len(added), ', '.join(r['id'] for r in added)))
for r in added:
    print('   %s  %s' % (r['id'], r['title']))
print('updating twins: %s' % ', '.join(twin_updates))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

# ---- JSON ----
for r in added:
    recs.append(r)
for rid, fields in twin_updates.items():
    by_id[rid].update(fields)
recs.sort(key=lambda r: r['id'])
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('JSON: %d records' % len(recs))

# ---- workbook ----
wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

plan = {}
for r in added:
    plan[r['id']] = {'title': r['title'], 'description': r['description'], 'notes': r['notes']}
for rid, fields in twin_updates.items():
    plan[rid] = dict(fields)

edits = []
for rid, fields in plan.items():
    row = rows[rid]
    for f, v in fields.items():
        c = head[f]
        cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
        if re.sub(r'\s+', ' ', cur).strip() != re.sub(r'\s+', ' ', v).strip():
            edits.append((rid, f, row, c, v))
tmp = BOOK + '.tmp'
for rid, f, row, c, v in edits:
    ws.cell(row=row, column=c, value=v)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-restore')
os.replace(tmp, BOOK)
print('workbook: %d cells written, 0 collateral' % len(edits))

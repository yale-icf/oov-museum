"""
fix-periods.py — align `period` with `issueYear`.

The user filtered the search page to 17th Century and found two documents that did
not belong. A full audit found 26: every record whose period contradicts its own
issueYear. The worst were goetzmann0499, an 1837 Mexican bond filed under 17th
Century, and goetzmann0500, a 1729 Ostend Company contract filed under 19th.

⚠️ Century convention. The collection uses the COMMON one — 1800 belongs to the 19th
century, 1900 to the 20th — not strict reckoning, where 1801–1900 is the 19th.
Established from the data rather than assumed: goetzmann0534 and goetzmann0549 are
both 1800 and both tagged 19th Century, and 447 of 473 records agree with the common
convention against 445 for strict. The two 1900 records disagreed with each other
(0647 said 19th, 0920 said 20th); 0920 is right and 0647 is corrected here.

Every one of the 26 was checked against its own description where the year could be
argued — 0304 "issued at Flensburg in May 1999" (was 21st Century), 0647 "dated by
hand the 20th of December 1900" (was 19th), 0738 "dated at Belgrade 1/13 August
1895" (was 20th). All three descriptions confirm the issueYear, so the period was
the wrong field in each case.

period is Excel-backed and drives a search facet, so both the workbook and the JSON
are written, and filter-index.json must be regenerated afterwards.

  py scratchpad/fix-periods.py                 # dry run
  py scratchpad/fix-periods.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

def band(year):
    y = int(year)
    if y >= 2000: return '21st Century'
    if y >= 1900: return '20th Century'
    if y >= 1800: return '19th Century'
    if y >= 1700: return '18th Century'
    if y >= 1600: return '17th Century'
    return 'Pre-17th Century'

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

plan = {}
for r in recs:
    ys = [y for y in (r.get('issueYear') or []) if str(y).isdigit()]
    ps = r.get('period') or []
    if not ys or not ps:
        continue
    want = band(ys[0])
    if want not in ps:
        plan[r['id']] = (ys[0], list(ps), want)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
col = head['period']
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for rid, (year, old, want) in sorted(plan.items(), key=lambda kv: kv[1][0]):
    r = rows[rid]
    cur = '' if ws.cell(row=r, column=col).value is None else str(ws.cell(row=r, column=col).value)
    if norm(cur) != norm(', '.join(old)):
        sys.exit('%s: sheet period %r does not match the JSON %r' % (rid, cur, old))
    edits.append((rid, r, year, cur, want))

print('%d period cell(s) to correct\n' % len(edits))
for rid, r, year, old, want in edits:
    print('  %s  %-6s %-22s -> %s' % (rid, year, old, want))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, r, year, old, want in edits:
    ws.cell(row=r, column=col, value=want)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, 'period') for _, r, _, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-periods')
os.replace(tmp, BOOK)

for rid, r, year, old, want in edits:
    by_id[rid]['period'] = [want]
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; %d cells written to the workbook and the JSON' % len(edits))
print('NOW REGENERATE data/filter-index.json — period is a facet.')

"""
save-0729-0295.py

Writes the at-risk values for goetzmann0729 and goetzmann0295 into the workbook, so
the next excel_to_json.py run cannot revert them.

Both records hold work that went straight into data/museum-data.json and was never
synced back. The import overwrites from the sheet, so without this they regress:

  goetzmann0729  the merged 0730 description (48w in the sheet, 120w in the JSON),
                 plus title, notes, keywords -- and `creator` in the sheet reads
                 "Kingdom of Bulgaria", which is wrong: the bond's face is headed
                 КНЯЖЕСТВО БЪЛГАРИЯ, Principality, verified from the image
                 2026-08-27. Bulgaria did not become a kingdom until 1908.
  goetzmann0295  notes, where the sheet still carries No. 011666 / 100,000,000 lei
                 against the correct 011648 / 160,000,000 read off the master.

It writes the CURRENT JSON values, not the pending facts-first rewrites -- the point
is to preserve what exists, not to apply the rewrite early.

Safety: works on a copy, verifies every one of the 868 rows and 20 columns is
untouched except the intended cells, and only then swaps the copy in. Refuses to run
against an open workbook.

  py scratchpad/save-0729-0295.py                 # dry run
  py scratchpad/save-0729-0295.py --write
"""

import json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv

lock = '~$' + BOOK
if os.path.exists(lock):
    sys.exit('REFUSING: %s is open in Excel (%s). Close it first.' % (BOOK, lock))

recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}

# field -> sheet column. keywords are pipe-joined on the way in, matching excel_to_json's parse.
PLAN = {
    'goetzmann0729': ['title', 'description', 'notes', 'creator', 'keywords'],
    'goetzmann0295': ['notes'],
}

def value_for(rec, field):
    if field == 'keywords':
        return '|'.join(rec.get('keywords') or [])
    if field == 'creator' and rec['id'] == 'goetzmann0729':
        return 'Principality of Bulgaria'      # verified from the bond's face
    return rec.get(field) or ''

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
fn_col = head['filename']

rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=fn_col).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

edits = []
for rid, fields in PLAN.items():
    if rid not in rows:
        sys.exit('no row for ' + rid)
    for f in fields:
        col = head[f]
        old = ws.cell(row=rows[rid], column=col).value
        old = '' if old is None else str(old)
        new = value_for(recs[rid], f)
        if old.strip() == new.strip():
            print('  %s.%s already correct, skipping' % (rid, f))
            continue
        edits.append((rid, f, rows[rid], col, old, new))

print('\n%d cells to write:' % len(edits))
for rid, f, r, c, old, new in edits:
    print('  %s  row %d  %-11s  %r' % (rid, r, f, (old[:58] + '…') if len(old) > 58 else old or '(blank)'))
    print('  %s              %-11s  -> %r' % (' ' * len(rid), '', (new[:58] + '…') if len(new) > 58 else new))

if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)

# verify: every cell identical to the original except the intended ones
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if before.shape != after.shape:
    os.remove(tmp); sys.exit('ABORT: shape changed %s -> %s' % (before.shape, after.shape))
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = []
for i in range(len(before)):
    for col in before.columns:
        if before.at[i, col] != after.at[i, col] and (i, col) not in intended:
            bad.append((i + 2, col))
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s), first: %s' % (len(bad), bad[:3]))

changed = sum(1 for i in range(len(before)) for col in before.columns
              if before.at[i, col] != after.at[i, col])
print('\nverified: %d cells changed, %d intended, 0 collateral' % (changed, len(edits)))

shutil.copy2(BOOK, BOOK + '.bak-before-0729-0295')
os.replace(tmp, BOOK)
print('WRITTEN to %s  (previous file kept as %s.bak-before-0729-0295)' % (BOOK, BOOK))

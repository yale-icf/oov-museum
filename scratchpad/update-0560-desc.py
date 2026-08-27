"""
update-0560-desc.py

Adds what the images gave to goetzmann0560's label, in the sheet and the JSON together.

The label ended at the 1910 write-down. Reading the four leaves at full resolution
(2026-08-27) turned up two further facts, neither of them anywhere in the record:
the reverse is handstamped with two repayment distributions of five francs, on the
decisions of the general meetings of 30 May 1918 and 1 June 1921; and the coupon
sheet is still attached and unclipped, its forty-eight coupons backed with the
financial years 1900 to 1947.

Written to both so an import cannot revert it and no import is needed to see it.

  py scratchpad/update-0560-desc.py                 # dry run
  py scratchpad/update-0560-desc.py --write
"""
import io, json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

ANCHOR = ' The concession was granted to build the railway inland'
ADD = (' The reverse carries two handstamps recording distributions of five francs each on '
       'repayment of the shares, decided by the general meetings of 30 May 1918 and 1 June 1921. '
       'The coupon sheet remains attached and unclipped, its forty-eight coupons backed with the '
       'financial years 1900 through 1947.')

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == 'goetzmann0560')
old = rec['description']
if ADD.strip() in old:
    sys.exit('already applied')
if ANCHOR not in old:
    sys.exit('anchor sentence not found; the label has changed')
new = old.replace(ANCHOR, ADD + ANCHOR, 1)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == 'goetzmann0560.jpg')
cur = str(ws.cell(row=row, column=head['description']).value or '').strip()
if cur != old.strip():
    sys.exit('the sheet and the JSON disagree; resolve that first')

print('row %d, %d -> %d words\n\n%s' % (row, len(old.split()), len(new.split()), new))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
ws.cell(row=row, column=head['description'], value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) != (row - 2, 'description')]
if bad:
    os.remove(tmp); sys.exit('ABORT: unintended changes %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-0560desc')
os.replace(tmp, BOOK)

rec['description'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

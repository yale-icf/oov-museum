"""Which goetzmannNNNN ids are taken, reserved, or free — so new material does not collide."""
import glob, json, os, re

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
live = {r['id'] for r in recs}
leaves = {p['id'] for r in recs for p in (r.get('pages') or [])}
num = lambda i: int(i.replace('goetzmann', ''))

taken = {num(i) for i in live | leaves}

# ids that exist as rows in the workbook (kept even when the record was dropped)
from openpyxl import load_workbook
wb = load_workbook('oov_data_new_edit_2.xlsx', read_only=True)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
rows = list(ws.values)
sheet = set()
for r in rows[1:]:
    v = r[1]
    if v and str(v).strip().lower().endswith('.jpg') and str(v).strip().lower().startswith('goetzmann'):
        try: sheet.add(num(str(v).strip()[:-4].lower()))
        except ValueError: pass

# masters on disk — anything photographed, whether catalogued or not
BASE = 'C:/Users/ks2479/Documents/my-project/origins-of-value/JPEG Files'
masters = set()
for p in glob.glob(BASE + '/**/*.jpg', recursive=True):
    m = re.match(r'goetzmann(\d{4})$', os.path.basename(p).lower()[:-4])
    if m: masters.add(int(m.group(1)))

used = taken | sheet | masters
print('ids in use somewhere: %d' % len(used))
print('  live records + leaves : %d' % len(taken))
print('  rows in the workbook  : %d' % len(sheet))
print('  masters photographed  : %d' % len(masters))
print('  highest id anywhere   : goetzmann%04d' % max(used))

# free gaps below the ceiling, and the clear space above
gaps, start = [], None
for i in range(1, max(used) + 1):
    if i not in used:
        start = start if start is not None else i
    elif start is not None:
        gaps.append((start, i - 1)); start = None
if start is not None:
    gaps.append((start, max(used)))
big = [g for g in gaps if g[1] - g[0] >= 2]
print('\ngaps of 3+ inside the numbered range (do NOT reuse — see below): %d' % len(big))
for a, b in big[:8]:
    print('   %04d-%04d  (%d ids)' % (a, b, b - a + 1))
print('\nCLEAR SPACE: goetzmann%04d and upward is unused everywhere.' % (max(used) + 1))

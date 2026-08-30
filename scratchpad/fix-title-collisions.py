"""
fix-title-collisions.py — separate the two title collisions the user's shortening pass
reintroduced, 2026-08-29.

  0302 / 0988  "Banque Industrielle de Chine Share" — the same instrument from two
               different issues, 1913 (statutes of 15 March 1913, capital 45,000,000
               francs) and 1920 (capital 150,000,000). The years differ, so a trailing
               year separates them, which is the convention already used for the
               Reichsbanknoten and the B&O certificates: a date appears in a title only
               to disambiguate.

  0353 / 1029  "Ostend Company Share Subscription Receipt" — both Antwerp 1723, both a
               250-guilder first instalment on a 1,000-guilder share, both ordering the
               same cashier Cogels junior to receive it. A year cannot separate them.
               What does is the KIND of document: 0353 is a manuscript receipt of 16
               August, 1029 a printed engraving of 13 August whose endorsements record
               three further instalments through 1725, so it doubles as a running
               receipt.

Kept short, since the user's pass is shortening titles.

  py scratchpad/fix-title-collisions.py                 # dry run
  py scratchpad/fix-title-collisions.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

PLAN = {
    'goetzmann0302': 'Banque Industrielle de Chine Share, 1913',
    'goetzmann0988': 'Banque Industrielle de Chine Share, 1920',
    'goetzmann0353': 'Manuscript Ostend Company Share Subscription Receipt',
    'goetzmann1029': 'Engraved Ostend Company Share Subscription Receipt',
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
col = head['title']
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for rid, new in PLAN.items():
    r = rows[rid]
    cur = '' if ws.cell(row=r, column=col).value is None else str(ws.cell(row=r, column=col).value)
    if norm(cur) != norm(by_id[rid].get('title')):
        sys.exit('%s: sheet and JSON disagree — %r vs %r' % (rid, cur, by_id[rid].get('title')))
    if norm(cur) != norm(new):
        edits.append((rid, r, cur, new))

print('%d title(s) to write' % len(edits))
for rid, r, old, new in edits:
    print('  %s  row %d\n     was: %s\n     now: %s' % (rid, r, old, new))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
for rid, r, old, new in edits:
    ws.cell(row=r, column=col, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, 'title') for _, r, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-collisions')
os.replace(tmp, BOOK)
for rid, r, old, new in edits:
    by_id[rid]['title'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

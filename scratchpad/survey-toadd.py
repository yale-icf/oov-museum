"""Survey To_Add_Files.xlsx: coverage against the tiled masters, and field completeness."""
import json, re
from collections import Counter
from openpyxl import load_workbook

wb = load_workbook('To_Add_Files.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c is not None else None for c in rows[0]]
idx = {h: i for i, h in enumerate(hdr) if h}
get = lambda r, k: (r[idx[k]] if k in idx and idx[k] < len(r) else None)

data = {}
for r in rows[1:]:
    f = get(r, 'Filename')
    if not f:
        continue
    k = str(f).strip()
    rid = (k[:-4] if k.lower().endswith('.jpg') else k).lower()
    data[rid] = r

src = json.load(open('scratchpad/new-batch-paths.json', encoding='utf-8'))
tiled = {'goetzmann' + n for n in src}
print('%d rows with a filename' % len(data))
print('  ids: %s … %s' % (min(data), max(data)))
print('  rows with no tiled master : %s' % (sorted(set(data) - tiled) or 'none'))
print('  tiled masters with no row : %s' % (sorted(tiled - set(data)) or 'none'))

fields = ['Document Title', 'Description', 'Type', 'Period', 'Location', 'Keywords',
          'Owner', 'Number of Pages', 'Item ID', 'Information']
print('\nfield completeness over %d rows:' % len(data))
for f in fields:
    n = sum(1 for r in data.values()
            if get(r, f) not in (None, '') and str(get(r, f)).strip())
    print('   %-18s %3d' % (f, n))

print('\nType values used:')
for v, c in Counter(str(get(r, 'Type') or '').strip() for r in data.values()).most_common():
    print('   %-40s %d' % (v[:40] or '(blank)', c))
print('\nPeriod values used:')
for v, c in Counter(str(get(r, 'Period') or '').strip() for r in data.values()).most_common():
    print('   %-40s %d' % (v[:40] or '(blank)', c))

# Number of Pages hints at multi-leaf grouping
print('\nNumber of Pages values:')
for v, c in Counter(str(get(r, 'Number of Pages') or '').strip() for r in data.values()).most_common(8):
    print('   %-10s %d' % (v or '(blank)', c))

json.dump({k: [str(x) if x is not None else None for x in v] for k, v in data.items()},
          open('scratchpad/toadd-rows.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
json.dump(hdr, open('scratchpad/toadd-header.json', 'w', encoding='utf-8'))
print('\n-> scratchpad/toadd-rows.json')

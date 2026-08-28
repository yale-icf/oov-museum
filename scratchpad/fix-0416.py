"""
fix-0416.py

Two readings corrected on goetzmann0416, both verified from the reconstructed image
at 3309x4690 (scratchpad/reconstruct.js) on 2026-08-27:

  address   the certificate reads "SIÈGE SOCIAL: 5, Allées Paul-Riquet, BÉZIERS".
            The description and notes both said 1, and both dropped the hyphen.
  printer   the imprint reads "LITH. CADENAT FRÈRES. BÉZIERS". The description
            said "Cadenet Frères" -- a in the second syllable, not e.

description and notes are Excel-backed, so both the sheet and the JSON are written.

  py scratchpad/fix-0416.py                 # dry run
  py scratchpad/fix-0416.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

RID = 'goetzmann0416'
SUBS = [
    ('seated at 1, Allées Paul Riquet in Béziers', 'seated at 5, Allées Paul-Riquet in Béziers'),
    ('Lithographed by Cadenet Frères of Béziers.', 'Lithographed by Cadenat Frères of Béziers.'),
    ('1, Allées Paul Riquet.', '5, Allées Paul-Riquet.'),
]

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == RID)
plan = {}
for field in ('description', 'notes'):
    cur = rec.get(field) or ''
    new = cur
    for a, b in SUBS:
        new = new.replace(a, b)
    if new != cur:
        plan[field] = new

if not plan:
    sys.exit('nothing to change -- already corrected?')
for a, b in SUBS:
    if not any(b in v for v in plan.values()) and not any(b in (rec.get(f) or '') for f in ('description', 'notes')):
        sys.exit('substitution never matched: %r' % a)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == RID + '.jpg')
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for field, value in plan.items():
    c = head[field]
    cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
    if norm(cur) != norm(rec.get(field)):
        sys.exit('%s: the sheet and the JSON disagree; resolve that first' % field)
    edits.append((field, c, cur, value))

print('%s row %d — %d cell(s)' % (RID, row, len(edits)))
for f, c, old, new in edits:
    for a, b in SUBS:
        if a in old:
            print('  %-11s %r -> %r' % (f, a, b))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for f, c, old, new in edits:
    ws.cell(row=row, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(row - 2, f) for f, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-0416')
os.replace(tmp, BOOK)

for field, value in plan.items():
    rec[field] = value
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

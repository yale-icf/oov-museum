"""
fix-dates-batch.py

Description-side date errors found by reading the images, 2026-08-27. These are the
cases where the DESCRIPTION was wrong -- the minority. In twelve other conflicts the
description was right and the transcription carried the error.

  0924  "issued in 1900" -> 1903. The red seal reads 明治三十六年四月二十五日,
        Meiji 36, 25 April 1903. issueYear 1900 -> 1903.
  0900  "30 July 1923" -> 20 July 1923. The sheet reads "Hagenow, den 20. Juli 1923".
        The year is right: redemption running from 1 July 1924 rules out 1925.
  0495  "Victor Hendrik Hoogenbergh" -> Pieter. Both the opening line and the
        signature read Pieter. The 1798 date in the description is correct; the
        transcription's 1778 is the error there.

description and notes are Excel-backed, so both the sheet and the JSON are written.

  py scratchpad/fix-dates-batch.py                 # dry run
  py scratchpad/fix-dates-batch.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

# id -> field -> (find, replace)
SUBS = {
    'goetzmann0924': {
        'description': [('a Japanese oil cooperative, issued in 1900.',
                         'a Japanese oil cooperative, issued 25 April 1903.')],
    },
    'goetzmann0900': {
        'description': [('Dated Hagenow, 30 July 1923', 'Dated Hagenow, 20 July 1923')],
    },
    'goetzmann0495': {
        'description': [('Victor Hendrik Hoogenbergh', 'Pieter Hendrik Hoogenbergh')],
    },
}
YEARS = {'goetzmann0924': ('1900', '1903', '25 April 1903')}   # id -> (old, new, issueDate cell)

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits, newvals = [], {}
for rid, fields in SUBS.items():
    rec, row = by_id[rid], rows[rid]
    for field, pairs in fields.items():
        cur_json = rec.get(field) or ''
        new = cur_json
        for a, b in pairs:
            if a not in new:
                sys.exit('%s.%s: pattern not found: %r' % (rid, field, a))
            new = new.replace(a, b)
        c = head[field]
        cur_sheet = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
        if norm(cur_sheet) != norm(cur_json):
            sys.exit('%s.%s: sheet and JSON disagree' % (rid, field))
        edits.append((rid, field, row, c, cur_sheet, new))
        newvals.setdefault(rid, {})[field] = new

for rid, (old, new, cell) in YEARS.items():
    row, c = rows[rid], head['issueDate']
    cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
    if norm(cur) != norm(cell):
        edits.append((rid, 'issueDate', row, c, cur, cell))

print('%d cell(s) to write' % len(edits))
for rid, f, r, c, old, new in edits:
    print('  %s %-12s row %d' % (rid, f, r))
    print('       was: ' + norm(old)[:130])
    print('       now: ' + norm(new)[:130])
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-datebatch')
os.replace(tmp, BOOK)

for rid, fields in newvals.items():
    for field, value in fields.items():
        by_id[rid][field] = value
by_id['goetzmann0924']['issueYear'] = ['1903']
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON (0924 issueYear 1900 -> 1903)')

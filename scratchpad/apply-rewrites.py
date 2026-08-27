"""
apply-rewrites.py

Writes the facts-first labels into the workbook, so the next excel_to_json.py run
carries them into data/museum-data.json instead of reverting them.

Three sets of cells, in one verified write:

  description  the 356 rewritten labels (sheet rows 382-869). Every one of these
               rows was checked to still hold the JSON text verbatim, so none of
               the user's own hand-edited descriptions is overwritten -- see
               scratchpad/rewrite-overlap-check.py.

  notes        wherever the JSON differs from the sheet. This is where the
               duplicate markers and the off-the-website statements live, and it
               is the same drift that would have reverted 0729 and 0295. The
               rewrite deliberately strips cataloguing asides out of the visitor-
               facing description, so notes has to be the place they survive:
                 goetzmann0494  a faded copy of this debenture was catalogued as
                                goetzmann0493 and dropped from the website
                 goetzmann0595  Qing-dynasty Chinese, not the Japanese/yen
                                attribution of earlier cataloguing
               (goetzmann0933's "formerly bound with goetzmann0931" already reads
               the same in both, so it needs no write.)

Anything found in notes is written BEFORE the description cells in the same pass,
so the sheet is never left holding a stripped description with no note behind it.

Safety: works on a copy, verifies all 868 rows x 20 columns are untouched except
the intended cells, then swaps the copy in and keeps a .bak. Refuses to run
against a workbook that is open in Excel.

  py scratchpad/apply-rewrites.py                 # dry run
  py scratchpad/apply-rewrites.py --write
"""

import json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv

lock = '~$' + BOOK
if os.path.exists(lock):
    sys.exit('REFUSING: %s is open in Excel (%s). Close it first.' % (BOOK, lock))

recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}
rew = json.load(open('scratchpad/rewrites.json', encoding='utf-8'))

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()
cell = lambda r, c: '' if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value)

edits = []

# notes first -- the asides have to have somewhere to live before the descriptions go
for rid, rec in sorted(recs.items()):
    if rid not in rows:
        continue
    new = rec.get('notes') or ''
    if norm(cell(rows[rid], head['notes'])) != norm(new):
        edits.append((rid, 'notes', rows[rid], head['notes'], cell(rows[rid], head['notes']), new))

guard = []
for rid, new in sorted(rew.items()):
    if rid not in rows:
        sys.exit('no row for ' + rid)
    r, c = rows[rid], head['description']
    # never overwrite a description the user has edited away from the JSON text
    if norm(cell(r, c)) != norm(recs[rid].get('description')):
        guard.append(rid); continue
    if norm(cell(r, c)) != norm(new):
        edits.append((rid, 'description', r, c, cell(r, c), new))

if guard:
    print('SKIPPED %d hand-edited row(s): %s' % (len(guard), ', '.join(guard)))

nn = sum(1 for e in edits if e[1] == 'notes')
print('%d cells to write: %d notes, %d descriptions' % (len(edits), nn, len(edits) - nn))
for rid, f, r, c, old, new in edits:
    if f == 'notes':
        print('  notes  %s  row %d' % (rid, r))
        print('     was: ' + (norm(old)[:110] or '(blank)'))
        print('     now: ' + norm(new)[:110])

before_words = sum(len((e[4] or '').split()) for e in edits if e[1] == 'description')
after_words = sum(len((e[5] or '').split()) for e in edits if e[1] == 'description')
print('\ndescriptions: %d -> %d words (%.0f%% cut)'
      % (before_words, after_words, 100 * (1 - after_words / before_words)))

if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)

before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if before.shape != after.shape:
    os.remove(tmp); sys.exit('ABORT: shape changed %s -> %s' % (before.shape, after.shape))
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s), first: %s' % (len(bad), bad[:3]))

changed = sum(1 for i in range(len(before)) for col in before.columns
              if before.at[i, col] != after.at[i, col])
print('\nverified: %d cells changed, %d intended, 0 collateral' % (changed, len(edits)))

shutil.copy2(BOOK, BOOK + '.bak-before-rewrites')
os.replace(tmp, BOOK)
print('WRITTEN to %s  (previous file kept as %s.bak-before-rewrites)' % (BOOK, BOOK))

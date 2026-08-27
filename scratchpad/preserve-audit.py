"""
preserve-audit.py -- run BEFORE applying the 356 rewritten labels.

Two ways the duplicate / not-on-the-website information could be lost:

  A. notes drift. The duplicate markers were written straight into
     data/museum-data.json (mark-duplicates.js). Where the sheet's notes cell
     still holds the old text, the next excel_to_json.py run reverts it -- the
     exact 0729/0295 failure, generalised to all 483.

  B. the rewrite dropped it. If a CURRENT description carried a duplicate or
     off-the-website statement and the new label does not, that fact has to land
     in notes before the description is overwritten.

Reports only. Writes nothing.
"""
import json, os, re, sys
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}
rew  = json.load(open('scratchpad/rewrites.json', encoding='utf-8'))

wb = load_workbook(BOOK, read_only=True)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
rows = list(ws.values)
head = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
fn, nc, dc = head['filename'], head['notes'], head['description']

sheet = {}
for r in rows[1:]:
    v = r[fn]
    if v and str(v).strip().endswith('.jpg'):
        sheet[str(v).strip()[:-4]] = {
            'notes': (r[nc] or '').strip(),
            'description': (r[dc] or '').strip(),
        }

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

PAT = re.compile(
    r'duplicat|\bdropped\b|design variant|same (?:engraved )?plate|another copy|'
    r'a (?:poorer|faded|second) copy|copy of this|goetzmann\s?\d{3,4}|'
    r'kept in the workbook|preserve numbering|catalogu|'
    r'not (?:shown|displayed|included|on) the (?:website|site)|from the website', re.I)

print('=== A. notes: JSON differs from the sheet (import would revert the JSON) ===\n')
a = []
for rid, rec in sorted(recs.items()):
    if rid not in sheet:
        continue
    j, s = norm(rec.get('notes')), norm(sheet[rid]['notes'])
    if j != s:
        a.append((rid, j, s, bool(PAT.search(j))))
for rid, j, s, dup in a:
    print(('  ** DUP-BEARING ' if dup else '     ') + rid)
    print('       sheet: ' + (s[:150] + '…' if len(s) > 150 else s or '(blank)'))
    print('       json : ' + (j[:150] + '…' if len(j) > 150 else j or '(blank)'))
print('\n  %d of 483 diverge; %d of those carry duplicate/off-site language.\n'
      % (len(a), sum(1 for x in a if x[3])))

print('=== B. rewrite dropped duplicate/off-site language from the description ===\n')
b = []
for rid, new in sorted(rew.items()):
    old = recs.get(rid, {}).get('description') or ''
    hits = set(m.group(0).lower() for m in PAT.finditer(old))
    if not hits:
        continue
    kept = set(m.group(0).lower() for m in PAT.finditer(new))
    lost = hits - kept
    if lost:
        b.append((rid, sorted(lost), old, new))
for rid, lost, old, new in b:
    print('  ' + rid + '   lost: ' + ', '.join(lost))
    for m in PAT.finditer(old):
        i = max(0, m.start() - 90)
        print('       …' + norm(old[i:m.end() + 90]) + '…')
    print('       notes now: ' + (norm(recs[rid].get('notes'))[:120] or '(blank)'))
print('\n  %d records.\n' % len(b))

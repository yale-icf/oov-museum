"""Record the image swap on goetzmann1027, in the workbook and the JSON."""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

RID = 'goetzmann1027'
ADD = (' Image replaced 2026-09-02 with the better of the collection\'s two scans of this same '
       'document, the one catalogued as goetzmann1031: 4539x3074 against 2560x1920, from a 2.1 MB '
       'master against 0.4 MB. The two are one physical document photographed twice (same '
       '"1627 11 Genn." annotation, same "quattro e mezzo" luoghi, same "quattrocento cinquanta" '
       'scudi, same signatures, same wax seal in the same position), so goetzmann1031 stays off '
       'the site as a duplicate scan rather than a second object.')

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == RID)
cur = (rec.get('notes') or '').rstrip()
if ADD.strip() in cur:
    sys.exit('already recorded')
new = (cur + ADD).strip()

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == RID + '.jpg')
sheet_now = '' if ws.cell(row=row, column=head['notes']).value is None else str(ws.cell(row=row, column=head['notes']).value)
if re.sub(r'\s+', ' ', sheet_now).strip() != re.sub(r'\s+', ' ', cur).strip():
    sys.exit('sheet and JSON notes disagree; resolve first')

print('%s row %d\n   now: %s' % (RID, row, new[-190:]))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
ws.cell(row=row, column=head['notes'], value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) != (row - 2, 'notes')]
if bad:
    os.remove(tmp); sys.exit('ABORT: %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-1027note')
os.replace(tmp, BOOK)
rec['notes'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

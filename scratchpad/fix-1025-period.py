"""
fix-1025-period.py

goetzmann1025, the Great Ming Circulating Treasure Note of 1375, is coded
period "17th Century". It is 14th-century, and the site already says so:
about.html gives the collection as spanning "14th through 20th centuries" with
1375 as the earliest object.

This surfaced while dropping goetzmann0393. That record was the only one coded
"Pre-17th Century", so removing it would have deleted the whole period band from
the facets and left 1375 filed under the 17th century -- the one record that
should have been holding that band up.

  py scratchpad/fix-1025-period.py                 # dry run
  py scratchpad/fix-1025-period.py --write
"""
import io, json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

RID, OLD, NEW = 'goetzmann1025', '17th Century', 'Pre-17th Century'

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == RID)
assert (rec.get('issueYear') or [''])[0] == '1375', rec.get('issueYear')
if rec.get('period') != [OLD]:
    sys.exit('period is already %s' % rec.get('period'))

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == RID + '.jpg')
cur = str(ws.cell(row=row, column=head['period']).value or '').strip()
if cur != OLD:
    sys.exit('sheet period is %r, expected %r' % (cur, OLD))

print('%s (issueYear 1375)  row %d  period %r -> %r' % (RID, row, OLD, NEW))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
ws.cell(row=row, column=head['period'], value=NEW)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) != (row - 2, 'period')]
if bad:
    os.remove(tmp); sys.exit('ABORT: unintended changes %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-1025period')
os.replace(tmp, BOOK)
rec['period'] = [NEW]
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('verified 0 collateral; written to the workbook and the JSON')

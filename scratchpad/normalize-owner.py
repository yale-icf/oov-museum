"""
normalize-owner.py — collapse the four Beinecke provenance variants to plain "Beinecke".

Per the user, 2026-08-28. The variants all describe how a piece reached the Beinecke;
the owner field records who holds it, so they are one value:

  Beinecke (Purchased by ICF transferred to Beinecke)        10
  Part of DLJ Collection - Joint purchase with ICF/Beinecke   9
  Joint ICF Beinecke Purchase                                 3
  Purchased by ICF transferred to Beinecke in 2003            1

"ICF" is left alone -- those are held by the ICF, not the Beinecke.

Nothing is lost: the acquisition detail survives in virtual_museum_database_full.xlsx
(its Owner and Information columns), which is where provenance narrative belongs.

  py scratchpad/normalize-owner.py            # dry run
  py scratchpad/normalize-owner.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel. Close it first.' % BOOK)

COLLAPSE = {
    'Beinecke (Purchased by ICF transferred to Beinecke)': 'Beinecke',
    'Part of DLJ Collection - Joint purchase with ICF/Beinecke': 'Beinecke',
    'Joint ICF Beinecke Purchase': 'Beinecke',
    'Purchased by ICF transferred to Beinecke in 2003': 'Beinecke',
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
col = head['owner']

edits = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if not v or not str(v).strip().endswith('.jpg'):
        continue
    rid = str(v).strip()[:-4]
    cur = '' if ws.cell(row=r, column=col).value is None else str(ws.cell(row=r, column=col).value).strip()
    if cur in COLLAPSE:
        edits.append((rid, r, cur, COLLAPSE[cur]))

from collections import Counter
print('%d owner cell(s) to collapse:' % len(edits))
for v, k in Counter(e[2] for e in edits).most_common():
    print('  %3d  %s  ->  Beinecke' % (k, v))
unseen = set(COLLAPSE) - {e[2] for e in edits}
if unseen:
    print('  (not found in the workbook: %s)' % ', '.join(sorted(unseen)))

if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, r, old, new in edits:
    ws.cell(row=r, column=col, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, 'owner') for _, r, _, _ in edits}
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-ownernorm')
os.replace(tmp, BOOK)

for rid, r, old, new in edits:
    if rid in by_id:
        by_id[rid]['owner'] = new
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; %d cells collapsed in the workbook and the JSON' % len(edits))

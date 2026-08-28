"""
merge-0386-into-0669.py

goetzmann0386 and goetzmann0669 are the same instrument: 1,000-mark bearer bonds of
the Russian 4% State Loan of 1902, same denomination, same equated values, differing
only in serial (No. 116485 against No. 196422). One is being dropped from the site.

0669 is kept because it is the better-documented record -- creator names the debt
administration, keywords are populated, notes carry the total issue, the description
runs 127 words against 94, and its currency uses spelled-out names rather than the
bare ISO code "DEM".

0386 was right on one field, so that is folded in first: language Russian + German.
0669's own transcription shows a German revenue stamp reading "REICHS-STEM[PEL]", so
German genuinely appears on the document and 0669's Russian-only value was wrong.

Its notes also gain the dropped id and serial, the convention used for 0295, 0494
and 1031.

  py scratchpad/merge-0386-into-0669.py                 # dry run
  py scratchpad/merge-0386-into-0669.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

KEEP, DROP = 'goetzmann0669', 'goetzmann0386'
recs = json.load(open('data/museum-data.json', encoding='utf-8'))
keep = next(r for r in recs if r['id'] == KEEP)
drop = next(r for r in recs if r['id'] == DROP)

NEW_LANG = ['Russian', 'German']
NEW_NOTES = (keep.get('notes') or '').rstrip()
addition = (' A second bond of the same issue and denomination, No. %s, was catalogued as %s and '
            'dropped from the website on 2026-08-27 as an indistinguishable duplicate; its row is '
            'kept in the workbook to preserve numbering.'
            % (next((v.replace('No. ', '') for v in (drop.get('identifiers') or []) if v.startswith('No.')), '116485'), DROP))
if addition.strip() not in NEW_NOTES:
    NEW_NOTES = (NEW_NOTES + addition).strip()

plan = {'language': ', '.join(NEW_LANG), 'notes': NEW_NOTES}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == KEEP + '.jpg')
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for field, value in plan.items():
    c = head[field]
    cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
    if norm(cur) != norm(value):
        edits.append((field, c, cur, value))

print('%s row %d — %d cell(s) to write' % (KEEP, row, len(edits)))
for f, c, old, new in edits:
    print('  %-9s was: %s' % (f, norm(old)[:120] or '(blank)'))
    print('  %-9s now: %s' % ('', norm(new)[:120]))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for f, c, old, new in edits:
    ws.cell(row=row, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(row - 2, f) for f, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-0386merge')
os.replace(tmp, BOOK)

keep['language'] = NEW_LANG
keep['notes'] = NEW_NOTES
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

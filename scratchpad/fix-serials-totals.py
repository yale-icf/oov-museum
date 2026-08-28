"""
fix-serials-totals.py — the 4 serial and 3 total conflicts, read from the images
2026-08-28.

Five of the seven were false positives in the checker, which matches any "No." and
any large number without knowing what it is:

  0668, 0676  the transcription's huge figures are the SAME total restated in other
              currencies, printed on the face: 310,498,000 gold rubles =
              1,241,992,000 francs (exactly 4x) = 1,003,529,536 marks, and so on.
  0450        the description's "No. 22" is a street address, Rue Caumartin No. 22.
              The share number, No. 1,735, is already in identifiers.
  0576        the description's "No. 4131" is the Peruvian law authorising the loan.
              The bond's serial, No. 71547, is already in identifiers.
  0485        a nominee register legitimately carries many numbers; No. 74 in the
              margin and No. 30 in the list are different things, not a conflict.

Two were real, and both were the description or notes rather than the transcription:

  0509  dated 1 January **1768**, not 1760. Three independent confirmations: the
        signature reads "17" printed with a handwritten 68; the printed terms run
        interest from "primo January 1700 Agt en Zestig" (1768) with the first due
        date "Negen en Zestig" (1769); and the coupon renewals endorsed on the sheet
        run in decades back to 1769. Its notes also said bond No. 99 (the sheet reads
        No. 90, as the description does) and "Capital ~500 Livres" (the sheet reads
        "Zes Honderd Guldens Courant", six hundred guilders, as the description does).
  0375  description RIGHT: the banner reads "EIGHT PER CENT FUND OF $1,500,000 /
        CREATED BY ACT OF CONGRESS FEB. 5TH 1840", repeated on all ten coupons. The
        transcription's "$5,000,000" is wrong. The bond is No. 1995, and identifiers
        was empty.

  py scratchpad/fix-serials-totals.py                 # dry run
  py scratchpad/fix-serials-totals.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

SHEET_SUBS = {
    'goetzmann0509': {
        'description': [('dated at Middelburg the first of January 1760',
                         'dated at Middelburg the first of January 1768')],
        'notes': [('Middelburg plantation bond No. 99.', 'Middelburg plantation bond No. 90.'),
                  ('Capital ~500 Livres at 5% p.a. January 1760.',
                   'Capital 600 guilders courant at 5% p.a. 1 January 1768.')],
        'issueDate': [('1760-01-01', '1 January 1768')],
        'identifiers': [('No. 99 | No. 90', 'No. 90')],
    },
    'goetzmann0375': {
        'identifiers': [('', 'No. 1995')],
    },
}
YEARS = {'goetzmann0509': '1768'}
TRANSCRIPTION = {
    'goetzmann0375': [(r'\$\s?5,000,000', '$1,500,000')],
    'goetzmann0485': [(r'No\.\s*50\.\s*Juffvr\. Henriette', 'No. 30. Juffvr. Henriette')],
}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits, newvals = [], {}
for rid, fields in SHEET_SUBS.items():
    row = rows[rid]
    for field, pairs in fields.items():
        c = head[field]
        cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
        new = cur
        for a, b in pairs:
            if a == '':
                if cur.strip():
                    sys.exit('%s.%s expected blank, found %r' % (rid, field, cur[:60]))
                new = b
            else:
                if a not in new:
                    sys.exit('%s.%s: pattern not found: %r (cell %r)' % (rid, field, a, cur[:90]))
                new = new.replace(a, b)
        if norm(cur) != norm(new):
            edits.append((rid, field, row, c, cur, new))
            newvals.setdefault(rid, {})[field] = new

print('%d cell(s) to write' % len(edits))
for rid, f, r, c, old, new in edits:
    print('  %s %-12s row %d' % (rid, f, r))
    print('       was: ' + (norm(old)[:130] or '(blank)'))
    print('       now: ' + norm(new)[:130])
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-serials')
os.replace(tmp, BOOK)

for rid, fields in newvals.items():
    for field, value in fields.items():
        by_id[rid][field] = [v.strip() for v in value.split('|')] if field == 'identifiers' else value
for rid, y in YEARS.items():
    by_id[rid]['issueYear'] = [y]
for rid, pairs in TRANSCRIPTION.items():
    t = by_id[rid].get('transcription') or ''
    for a, b in pairs:
        n = len(re.findall(a, t))
        t = re.sub(a, b, t)
        print('  transcription %s: %d replacement(s) for %s' % (rid, n, a))
    by_id[rid]['transcription'] = t
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON (0509 issueYear 1760 -> 1768)')

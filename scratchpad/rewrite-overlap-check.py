"""Would applying the 356 labels overwrite anything the user edited by hand?

For each rewrite target, compare the sheet's description with the JSON's. They
should be identical: the rewrites cover only rows the user had NOT touched. Any
divergence is a hand edit that the label would clobber, and must be reviewed.
"""
import json, re
from openpyxl import load_workbook

recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}
rew  = json.load(open('scratchpad/rewrites.json', encoding='utf-8'))
wb = load_workbook('oov_data_new_edit_2.xlsx', read_only=True)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
rows = list(ws.values)
head = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
fn, dc = head['filename'], head['description']

sheet, rownum = {}, {}
for n, r in enumerate(rows[1:], start=2):
    v = r[fn]
    if v and str(v).strip().endswith('.jpg'):
        sheet[str(v).strip()[:-4]] = (r[dc] or '').strip()
        rownum[str(v).strip()[:-4]] = n

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()
miss, diff = [], []
for rid in sorted(rew):
    if rid not in sheet:
        miss.append(rid); continue
    if norm(sheet[rid]) != norm(recs[rid].get('description')):
        diff.append(rid)

print('%d labels; %d have no row in the sheet; %d have a sheet description that '
      'differs from the JSON' % (len(rew), len(miss), len(diff)))
if miss: print('  no row: ' + ', '.join(miss))
for rid in diff:
    print('\n  %s  (row %d)' % (rid, rownum[rid]))
    print('    sheet: ' + norm(sheet[rid])[:200])
    print('    json : ' + norm(recs[rid].get('description'))[:200])
rn = [rownum[r] for r in rew if r in rownum]
print('\nrow range touched: %d–%d' % (min(rn), max(rn)))

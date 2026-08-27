"""
fix-issuedates.py

Six issueYear corrections were applied to data/museum-data.json but never to the
workbook. excel_to_json.py derives issueYear from the sheet's `issueDate` column,
so without these the next import reverts all six.

  0640  1965 -> 1990   Belgrade, February 1990 on its face; coupons 1996-98
  0641  1970 -> 1987   signed Sarajevo 1 October 1987; annuities 1990-93
  0638  1928 -> 1914   the loan names itself "of 1914"; Governor Potiorek, in post to 1914
  0646  1927 -> 1930   dated 1 September 1930 at Ujpest
  0631  1832 -> 1825   the sheet's 1832 under a prospectus of 1834 dated the certificate
                       two years before its own prospectus; description, transcription and
                       siblings 0525/0616 all give 1825 under a prospectus of 1824
  0343  1920 -> 1928   read from the image: the certificate carries NO date at all, only a
                       printed notice that stamp duty was discharged under an authorisation
                       published in the Journal Officiel of 3 April 1928. Written as
                       "ca. 1928" because that is an inference, not a date on the document.

0631's notes carry the same impossible pair and are corrected with it; the script
syncs notes wherever the JSON and the sheet disagree, so that is picked up here.

Safety: works on a copy, verifies every other cell is untouched, keeps a .bak,
refuses to run against an open workbook.

  py scratchpad/fix-issuedates.py                 # dry run
  py scratchpad/fix-issuedates.py --write
"""

import json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv

lock = '~$' + BOOK
if os.path.exists(lock):
    sys.exit('REFUSING: %s is open in Excel (%s). Close it first.' % (BOOK, lock))

DATES = {
    'goetzmann0640': 'February 1990',
    'goetzmann0641': '1 October 1987',
    'goetzmann0638': '1 April 1914',
    'goetzmann0646': '1 September 1930',
    'goetzmann0631': '31 December 1825',
    'goetzmann0343': 'ca. 1928',
}

recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()
cell = lambda r, c: '' if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value)

edits = []
for rid, new in DATES.items():
    if rid not in rows:
        sys.exit('no row for ' + rid)
    r, c = rows[rid], head['issueDate']
    want = (recs[rid].get('issueYear') or [''])[0]
    got = re.search(r'\b(1[0-9]{3}|20[0-9]{2})\b', new)
    if not got or got.group(1) != want:
        sys.exit('%s: %r would not yield issueYear %s' % (rid, new, want))
    if norm(cell(r, c)) != norm(new):
        edits.append((rid, 'issueDate', r, c, cell(r, c), new))

for rid, rec in sorted(recs.items()):
    if rid not in rows:
        continue
    r, c = rows[rid], head['notes']
    if norm(cell(r, c)) != norm(rec.get('notes')):
        edits.append((rid, 'notes', r, c, cell(r, c), rec.get('notes') or ''))

print('%d cells to write' % len(edits))
for rid, f, r, c, old, new in edits:
    print('  %-11s %s  row %d' % (f, rid, r))
    print('     was: ' + (norm(old)[:130] or '(blank)'))
    print('     now: ' + norm(new)[:130])

if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)

before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
if before.shape != after.shape:
    os.remove(tmp); sys.exit('ABORT: shape changed %s -> %s' % (before.shape, after.shape))
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s), first: %s' % (len(bad), bad[:3]))

changed = sum(1 for i in range(len(before)) for col in before.columns
              if before.at[i, col] != after.at[i, col])
print('\nverified: %d cells changed, %d intended, 0 collateral' % (changed, len(edits)))

shutil.copy2(BOOK, BOOK + '.bak-before-issuedates')
os.replace(tmp, BOOK)
print('WRITTEN to %s' % BOOK)

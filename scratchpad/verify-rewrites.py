"""Post-write verification: the labels landed, and every duplicate / off-the-website
note is present in the workbook (not just in the JSON)."""
import json, re
from openpyxl import load_workbook

recs = {r['id']: r for r in json.load(open('data/museum-data.json', encoding='utf-8'))}
rew = json.load(open('scratchpad/rewrites.json', encoding='utf-8'))
wb = load_workbook('oov_data_new_edit_2.xlsx', read_only=True)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
rows = list(ws.values)
head = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
fn, dc, nc = head['filename'], head['description'], head['notes']

sheet = {}
for r in rows[1:]:
    v = r[fn]
    if v and str(v).strip().endswith('.jpg'):
        sheet[str(v).strip()[:-4]] = ((r[dc] or '').strip(), (r[nc] or '').strip())

norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()
bad = [rid for rid, t in rew.items() if norm(sheet.get(rid, ('', ''))[0]) != norm(t)]
print('descriptions: %d of %d labels present verbatim in the sheet%s'
      % (len(rew) - len(bad), len(rew), '' if not bad else '  MISSING: ' + ', '.join(bad)))

drift = [rid for rid in recs if rid in sheet and norm(recs[rid].get('notes')) != norm(sheet[rid][1])]
print('notes: %d of %d records diverge between JSON and sheet' % (len(drift), len(recs)))
if drift: print('  ' + ', '.join(drift))

DUP = re.compile(r'duplicat|design variant|dropped (?:from|as)|another copy|'
                 r'a (?:poorer|faded|second) copy|kept in the workbook|preserve numbering|'
                 r'superseded|catalogued as goetzmann', re.I)
print('\nduplicate / off-the-website notes now carried by the workbook:')
for rid in sorted(sheet):
    n = sheet[rid][1]
    if DUP.search(n):
        print('  %s  %s' % (rid, norm(n)[:170] + ('…' if len(norm(n)) > 170 else '')))

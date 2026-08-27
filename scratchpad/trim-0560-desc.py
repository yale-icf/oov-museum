"""Tighten goetzmann0560's label after the addition. No fact removed -- the 1899
capital increase, the 1910 write-down, the 1918/1921 repayments, the coupon years,
the concession and the printer all remain. Cuts the restated "registered office in
Paris", the editorial "record what became of it" and "a five-fold write-down", and
the redundant "of 500 francs each"."""
import io, json, os, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

NEW = ("A fully paid bearer share of five hundred francs in the Compagnie Impériale des Chemins "
       "de Fer Éthiopiens, dated Paris, 14 December 1899. The company was constituted under "
       "statutes deposited on 7 August 1896 before Maître Rey, notary in Paris, where it kept its "
       "registered office; its capital of 8,000,000 francs was raised to 18,000,000 by the "
       "extraordinary general meeting of 12 December 1899 and divided into 36,000 shares of 500 "
       "francs. The title panel carries the company name in Amharic above the French. Two later "
       "overstamps reduce the capital to 3,600,000 francs by the general meeting of 4 June 1910 "
       "and restrike the share as one of one hundred francs. The reverse is handstamped with "
       "distributions of five francs each on repayment of the shares, decided by the general "
       "meetings of 30 May 1918 and 1 June 1921. The coupon sheet remains attached and unclipped, "
       "its forty-eight coupons backed with the financial years 1900 through 1947. The concession "
       "was granted to build the railway inland from the Gulf port of Djibouti. Printed by "
       "Imprimerie Chaix, 20 Rue Bergère, Paris.")

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == 'goetzmann0560')
old = rec['description']
for must in ['30 May 1918', '1 June 1921', '1900 through 1947', '4 June 1910', '14 December 1899',
             '7 August 1896', '18,000,000', '3,600,000', 'Djibouti', 'Chaix']:
    assert must in NEW, 'lost: ' + must

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == 'goetzmann0560.jpg')
if str(ws.cell(row=row, column=head['description']).value or '').strip() != old.strip():
    sys.exit('the sheet and the JSON disagree; resolve that first')

print('%d -> %d words\n\n%s' % (len(old.split()), len(NEW.split()), NEW))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
ws.cell(row=row, column=head['description'], value=NEW)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
bad = [(i + 2, c) for i in range(len(before)) for c in before.columns
       if before.at[i, c] != after.at[i, c] and (i, c) != (row - 2, 'description')]
if bad:
    os.remove(tmp); sys.exit('ABORT: unintended changes %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-0560trim')
os.replace(tmp, BOOK)
rec['description'] = NEW
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON')

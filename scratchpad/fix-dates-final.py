"""
fix-dates-final.py — the last of the 27 date conflicts, read from the images 2026-08-28.

goetzmann1034, Middelburg plantation bond. Dated **1 January 1771**, not 1794:
  - the signature line reads "Middelburg den Eersten Januarij 1771";
  - the printed terms run interest from "primo January 1700 Een en Zeventig" (1771)
    with the first due date "primo January 1700 twee en Zeventig" (1772);
  - the coupon renewals endorsed on the sheet run in decades back from 1799 à 1808,
    which lands on 1771 and cannot reach 1794.
  Its number is **No. 572**, not the No. 502 in the notes and identifiers.

goetzmann0683, Insull Utility Investments debenture. The description's warrant dates
were wrong and its notes described a different instrument:
  - the footnote reads "ONE VOID AFTER DECEMBER 31, 1930, AND THE OTHER VOID AFTER
    DECEMBER 31, 1934". The description said 1933 and 1934.
  - the plate reads "TEN-YEAR 6% GOLD DEBENTURE, SERIES B, DUE JANUARY 1, 1940".
    The notes said 7%, Series A, due January 1, 1949 -- three errors in one line.
  The description was right about the rate, series and maturity.

Two transcriptions are corrected with them: 0602 (1914 -> 1922; the bond attests "as
of the fifteenth day of July, One thousand nine hundred and twenty-two") and 0613
(1890 -> 1886; "Amsterdam, 8 December 1886"). ⚠️ 0613's OTHER year, 1880, is genuine --
the Administratie-Kantoor's notice of December 1880 -- and is left alone.

  py scratchpad/fix-dates-final.py                 # dry run
  py scratchpad/fix-dates-final.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

SHEET_SUBS = {
    'goetzmann1034': {
        'description': [('dated at Middelburg in January 1794', 'dated at Middelburg on 1 January 1771')],
        'notes': [('Plantation bond No. 502, Register Middelburg, January 1794',
                   'Plantation bond No. 572, Register Middelburg, 1 January 1771')],
        'issueDate': [('1794-01', '1 January 1771')],
        'identifiers': [('No. 502', 'No. 572')],
    },
    'goetzmann0683': {
        'description': [('voided after December 31, 1933 and December 31, 1934',
                         'voided after December 31, 1930 and December 31, 1934')],
        'notes': [('Insull Utility Investments, Inc., 7% Gold Debenture, Series A, Due January 1, 1949.',
                   'Insull Utility Investments, Inc., 6% Gold Debenture, Series B, due January 1, 1940.')],
    },
}
YEARS = {'goetzmann1034': '1771'}
TRANSCRIPTION = {'goetzmann0602': [('1914', '1922')], 'goetzmann0613': [('1890', '1886')],
                 'goetzmann0683': [('1946', '1940')]}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
by_id = {r['id']: r for r in recs}

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
rows = {}
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=head['filename']).value
    if v and str(v).strip().endswith('.jpg'):
        rows[str(v).strip()[:-4]] = r
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits, newvals = [], {}
for rid, fields in SHEET_SUBS.items():
    row = rows[rid]
    for field, pairs in fields.items():
        c = head[field]
        cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
        new = cur
        for a, b in pairs:
            if a not in new:
                sys.exit('%s.%s: pattern not found: %r (cell is %r)' % (rid, field, a, cur[:90]))
            new = new.replace(a, b)
        edits.append((rid, field, row, c, cur, new))
        newvals.setdefault(rid, {})[field] = new

print('%d cell(s) to write' % len(edits))
for rid, f, r, c, old, new in edits:
    print('  %s %-12s row %d' % (rid, f, r))
    print('       was: ' + norm(old)[:135])
    print('       now: ' + norm(new)[:135])
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
for rid, f, r, c, old, new in edits:
    ws.cell(row=r, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(r - 2, f) for _, f, r, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-datesfinal')
os.replace(tmp, BOOK)

for rid, fields in newvals.items():
    for field, value in fields.items():
        by_id[rid][field] = [value] if field == 'identifiers' else value
for rid, y in YEARS.items():
    by_id[rid]['issueYear'] = [y]
for rid, pairs in TRANSCRIPTION.items():
    t = by_id[rid].get('transcription') or ''
    n = 0
    for a, b in pairs:
        n += len(re.findall(r'\b%s\b' % a, t))
        t = re.sub(r'\b%s\b' % a, b, t)
    by_id[rid]['transcription'] = t
    print('  transcription %s: %d token(s) %s' % (rid, n, pairs))
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON (1034 issueYear 1794 -> 1771)')

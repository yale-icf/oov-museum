"""
fix-0568.py

goetzmann0568, the Costa Rica Railway £100 Second Debenture. The date was settled by
the user 2026-08-27: the year is written OVER the printed "188...", which is why the
strokes read ambiguously as 1880 / 1888 / 1890. It is 1890.

With the image open for that, two further errors showed up:

  rate     the plate reads "BEARING INTEREST AT THE RATE OF 6 PER CENT PER ANNUM".
           The description and notes both said five percent.
  printer  the imprint reads "DOHERTY & Co, 6, Gt. NEWPORT ST, W.C." The description
           said "Goertz & Co., 57 W.C." -- wrong name and wrong address.

The day, 28 February, is on the seal line and was missing from the description.

  py scratchpad/fix-0568.py                 # dry run
  py scratchpad/fix-0568.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

RID = 'goetzmann0568'
DESC = ("A £100 Second Debenture of the Costa Rica Railway Company, Limited, sealed under the "
        "company's common seal on 28 February 1890. The certificate forms part of a £600,000 "
        "issue of 6,000 such debentures, each bearing interest at six percent a year, payable "
        "half-yearly on the first of March and the first of September. It secures repayment of "
        "principal out of the company's revenues and is transferable on the register. Printed by "
        "Doherty & Co., 6 Great Newport Street, London W.C.")
NOTES = ("Second debenture No. 6886; £600,000 total issue; 6,000 bonds numbered 6,551–12,550; "
         "6% p.a.; portrait vignette; British company operating Costa Rican railway. Sealed "
         "28 February 1890 — the year is written over the printed \"188…\", which is why it reads "
         "ambiguously. Corrected from the image 2026-08-27: the record had said February 1888, "
         "five percent, and \"Goertz & Co., 57 W.C.\"")
PLAN = {'description': DESC, 'notes': NOTES, 'issueDate': '28 February 1890'}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == RID)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == RID + '.jpg')
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for field, value in PLAN.items():
    c = head[field]
    cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
    if norm(cur) != norm(value):
        edits.append((field, c, cur, value))

print('%s row %d — %d cell(s)' % (RID, row, len(edits)))
for f, c, old, new in edits:
    print('  %-12s was: %s' % (f, norm(old)[:130] or '(blank)'))
    print('  %-12s now: %s' % ('', norm(new)[:130]))
if not WRITE:
    print('\ndry run -- pass --write to apply'); sys.exit(0)

tmp = BOOK + '.tmp'
for f, c, old, new in edits:
    ws.cell(row=row, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(row - 2, f) for f, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %s' % bad[:3])
shutil.copy2(BOOK, BOOK + '.bak-before-0568')
os.replace(tmp, BOOK)

rec['description'] = DESC
rec['notes'] = NOTES
rec['issueYear'] = ['1890']
if rec.get('transcription'):
    rec['transcription'] = re.sub(r'\b1880\b', '1890', rec['transcription'])
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON (issueYear 1888 -> 1890)')

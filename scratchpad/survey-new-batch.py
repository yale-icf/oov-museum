"""Survey a new batch of masters before ingesting: naming, size, resolution, workbook rows."""
import glob, json, os, re, sys
from openpyxl import load_workbook

BASE = 'C:/Users/ks2479/Documents/my-project/origins-of-value/JPEG Files'
LO, HI = 1100, 1199

paths = {}
odd_case, odd_ext = [], []
for p in glob.glob(BASE + '/**/*', recursive=True):
    if not os.path.isfile(p):
        continue
    b = os.path.basename(p)
    m = re.match(r'goetzmann(\d{4})\.(jpg|jpeg|tif|tiff|png)$', b, re.I)
    if not m:
        continue
    n = int(m.group(1))
    if not (LO <= n <= HI):
        continue
    if b != b.lower():
        odd_case.append(b)
    if m.group(2).lower() != 'jpg':
        odd_ext.append(b)
    paths[n] = p

print('%d masters found in %d-%d' % (len(paths), LO, HI))
missing = [i for i in range(LO, HI + 1) if i not in paths]
print('  ids absent from the run : %s' % (missing if missing else 'none'))
print('  non-lowercase filenames : %s' % (odd_case if odd_case else 'none'))
print('  non-.jpg extensions     : %s' % (odd_ext if odd_ext else 'none'))

sizes = sorted((os.path.getsize(p), n) for n, p in paths.items())
tot = sum(s for s, _ in sizes)
print('\nfile size: total %.2f GB, mean %.1f MB, min %.1f MB (goetzmann%04d), max %.1f MB (goetzmann%04d)'
      % (tot / 1e9, tot / len(sizes) / 1e6, sizes[0][0] / 1e6, sizes[0][1],
         sizes[-1][0] / 1e6, sizes[-1][1]))
small = [(n, s) for s, n in sizes if s < 1_000_000]
print('  under 1 MB (would tile softly, as the ICF batch does): %d %s'
      % (len(small), [('goetzmann%04d' % n) for n, _ in small][:8]))

# workbook rows?
wb = load_workbook('oov_data_new_edit_2.xlsx', read_only=True)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
rows = list(ws.values)
head = {str(v).strip(): i for i, v in enumerate(rows[0]) if v}
present, filled = set(), {}
for r in rows[1:]:
    v = r[head['filename']]
    if not v:
        continue
    m = re.match(r'goetzmann(\d{4})\.jpg$', str(v).strip(), re.I)
    if m and LO <= int(m.group(1)) <= HI:
        n = int(m.group(1))
        present.add(n)
        for f in ('title', 'description', 'type', 'issueDate', 'period',
                  'issuingCountry', 'currency', 'language', 'owner'):
            if f in head and r[head[f]] not in (None, '') and str(r[head[f]]).strip():
                filled[f] = filled.get(f, 0) + 1
print('\nworkbook rows for this range: %d' % len(present))
if present:
    for f, c in sorted(filled.items(), key=lambda kv: -kv[1]):
        print('   %-16s %d filled' % (f, c))
else:
    print('   none yet — the data has not been entered')

json.dump({str(n): p for n, p in sorted(paths.items())},
          open('scratchpad/new-batch-paths.json', 'w', encoding='utf-8'), indent=2)
print('\n-> scratchpad/new-batch-paths.json')

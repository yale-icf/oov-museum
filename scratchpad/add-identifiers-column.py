"""
add-identifiers-column.py

Adds the internal `identifiers` column to an OOV workbook and fills it from
data/museum-data.json, so the seeded serial / loan / certificate numbers become
editable in Excel and survive future excel_to_json.py runs.

RUN THIS ONLY WITH THE WORKBOOK CLOSED. It refuses to touch a file that still has
an Excel lock (~$name.xlsx) beside it.

  py scratchpad/add-identifiers-column.py                       # dry run, oov_data_new.xlsx
  py scratchpad/add-identifiers-column.py --write
  py scratchpad/add-identifiers-column.py --file oov_data_new_edit.xlsx --write

The sheet is per-IMAGE (one row per filename) while identifiers are per-RECORD, so
values are written to the record's primary row only; sub-page rows are left blank.
Values are joined with " | " -- never a comma, which these numbers contain.
"""

import argparse
import json
import os
import sys

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT, 'data', 'museum-data.json')
HEADER = 'identifiers'
JOIN = ' | '


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default='oov_data_new.xlsx')
    ap.add_argument('--write', action='store_true')
    args = ap.parse_args()

    path = os.path.join(ROOT, args.file)
    lock = os.path.join(os.path.dirname(path), '~$' + os.path.basename(path))
    if os.path.exists(lock):
        sys.exit('REFUSING: %s is open in Excel (lock file %s). Close it first.'
                 % (args.file, os.path.basename(lock)))
    if not os.path.exists(path):
        sys.exit('no such workbook: ' + path)

    with open(JSON_PATH, encoding='utf-8') as f:
        records = json.load(f)
    by_id = {r['id']: r.get('identifiers', []) for r in records}

    wb = load_workbook(path)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    if 'filename' not in headers:
        sys.exit('no `filename` column in %s -- wrong sheet?' % args.file)

    col = headers.get(HEADER)
    fresh = col is None
    if fresh:
        col = ws.max_column + 1
        if args.write:
            ws.cell(row=1, column=col, value=HEADER)

    fn_col = headers['filename']
    filled = blank = missing = 0

    for r in range(2, ws.max_row + 1):
        fn = ws.cell(row=r, column=fn_col).value
        if not fn or not str(fn).strip().endswith('.jpg'):
            continue
        item_id = str(fn).strip()[:-4]
        if item_id not in by_id:
            missing += 1
            continue
        vals = by_id[item_id]
        if vals:
            filled += 1
            if args.write:
                ws.cell(row=r, column=col, value=JOIN.join(vals))
        else:
            blank += 1

    if args.write:
        wb.save(path)

    print('%s: column `%s` %s at index %d' % (args.file, HEADER,
                                              'added' if fresh else 'already present', col))
    print('  %d rows filled, %d left blank, %d rows have no matching record (sub-pages / dropped)'
          % (filled, blank, missing))
    print('  WRITTEN' if args.write else '  dry run -- pass --write to apply')


if __name__ == '__main__':
    main()

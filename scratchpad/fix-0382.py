"""
fix-0382.py

goetzmann0382, the Nippon Kangyō Bank premium savings bond, was wrong in four
places. All four read off the reconstructed image on 2026-08-27:

  date          the panel reads 昭和十五年十一月 -- Shōwa 15, eleventh month =
                November 1940. The border repeats 紀元二千六百年, imperial year
                2600, which is also 1940. The record said November 1938.
  issue number  the heading reads 第拾八回, the EIGHTEENTH issue. The record said
                the eighty-eighth. (第八拾八回 would be the 88th.)
  denomination  the face reads 金拾五圓, fifteen yen, against the ten-yen sale
                price below it (割引賣出價格金拾圓) -- a coherent discount. The
                description said fifty yen and the notes said fifty-five.
  serial        No. 069959 on the face; the sheet's identifiers said No. 009951.

The notes were worse than the description: they called the issuer the "Japan
Industrial Bank" (it is the Nippon Kangyō Bank, correct in the title) and dated it
Shōwa 35, 1960 -- twenty years out.

description, notes and issueDate are Excel-backed; identifiers is a sheet column
too. All are written to both the workbook and the JSON.

  py scratchpad/fix-0382.py                 # dry run
  py scratchpad/fix-0382.py --write
"""
import io, json, os, re, shutil, sys
import pandas as pd
from openpyxl import load_workbook

BOOK = 'oov_data_new_edit_2.xlsx'
WRITE = '--write' in sys.argv
if os.path.exists('~$' + BOOK):
    sys.exit('REFUSING: %s is open in Excel.' % BOOK)

RID = 'goetzmann0382'
DESC = ('Fifteen yen, the eighteenth issue of premium savings bonds sold by the Nippon Kangyō '
        'Bank, the government-linked Hypothec Bank of Japan, dated November 1940. The bank sold '
        'the bond to the public at a discounted price of ten yen; the gap between that price and '
        'the redemption value stood in place of interest. An attached prize drawing offered '
        'lottery-style bonuses to holders.')
NOTES = ('Nippon Kangyō Bank (日本勧業銀行) savings bond (貯蓄債券), 18th issue (第拾八回). '
         'Face 15 yen (金拾五圓); discount sale price 10 yen (割引賣出價格金拾圓). No. 069959. '
         'Dated 昭和十五年十一月 = November 1940, corroborated by the border 紀元二千六百年 '
         '(imperial year 2600 = 1940). Corrected from the image 2026-08-27: the record had read '
         '88th issue, 50/55 yen, No. 009951, "Japan Industrial Bank", and Shōwa 35 (1960).')
PLAN = {'description': DESC, 'notes': NOTES, 'issueDate': 'November 1940', 'identifiers': 'No. 069959'}

recs = json.load(open('data/museum-data.json', encoding='utf-8'))
rec = next(r for r in recs if r['id'] == RID)

wb = load_workbook(BOOK)
ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
head = {str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
row = next(r for r in range(2, ws.max_row + 1)
           if str(ws.cell(row=r, column=head['filename']).value).strip() == RID + '.jpg')
norm = lambda s: re.sub(r'\s+', ' ', (s or '')).strip()

edits = []
for field, value in PLAN.items():
    c = head[field]
    cur = '' if ws.cell(row=row, column=c).value is None else str(ws.cell(row=row, column=c).value)
    if norm(cur) != norm(value):
        edits.append((field, c, cur, value))

print('%s row %d — %d cell(s)' % (RID, row, len(edits)))
for f, c, old, new in edits:
    print('  %-12s was: %s' % (f, norm(old)[:120] or '(blank)'))
    print('  %-12s now: %s' % ('', norm(new)[:120]))
if not WRITE:
    print('\ndry run -- pass --write to apply')
    sys.exit(0)

tmp = BOOK + '.tmp'
for f, c, old, new in edits:
    ws.cell(row=row, column=c, value=new)
wb.save(tmp)
before = pd.read_excel(BOOK, dtype=str).fillna('')
after = pd.read_excel(tmp, dtype=str).fillna('')
intended = {(row - 2, f) for f, _, _, _ in edits}
bad = [(i + 2, col) for i in range(len(before)) for col in before.columns
       if before.at[i, col] != after.at[i, col] and (i, col) not in intended]
if bad:
    os.remove(tmp); sys.exit('ABORT: %d unintended change(s): %s' % (len(bad), bad[:3]))
shutil.copy2(BOOK, BOOK + '.bak-before-0382')
os.replace(tmp, BOOK)

rec['description'] = DESC
rec['notes'] = NOTES
rec['issueYear'] = ['1940']
rec['identifiers'] = ['No. 069959']
io.open('data/museum-data.json', 'w', encoding='utf-8', newline='\n').write(
    json.dumps(recs, ensure_ascii=False, indent=2) + '\n')
print('\nverified 0 collateral; written to the workbook and the JSON (issueYear 1938 -> 1940)')
